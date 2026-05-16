from openpyxl import load_workbook


class CampaignLoader:
    def __init__(self, workbook_path):
        self.workbook_path = workbook_path

    def _load_rows(self, sheet_name):
        wb = load_workbook(self.workbook_path, data_only=True)

        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Missing sheet: {sheet_name}")

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return []

        headers = [str(h).strip() if h else "" for h in rows[0]]
        data = []

        for row in rows[1:]:
            item = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                item[header] = row[idx] if idx < len(row) else None

            if any(v is not None and str(v).strip() != "" for v in item.values()):
                data.append(item)

        return data

    def load_campaign(self, campaign_id):
        rows = self._load_rows("Campaign_Control")

        for row in rows:
            if str(row.get("campaign_id", "")).strip() == str(campaign_id).strip():
                return row

        raise ValueError(f"Campaign not found: {campaign_id}")
