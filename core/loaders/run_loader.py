from openpyxl import load_workbook


class RunLoader:
    def __init__(self, workbook_path):
        self.workbook_path = workbook_path

    def _load_rows(self, sheet_name):
        wb = load_workbook(self.workbook_path, data_only=True)

        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Missing sheet: {sheet_name}")

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return []

        headers = [str(h).strip() if h else "" for h in rows[0]]
        data = []

        for row in rows[1:]:
            item = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                item[header] = row[idx] if idx < len(row) else None

            if any(v is not None and str(v).strip() != "" for v in item.values()):
                data.append(item)

        return data

    def get_pending_runs(self):
        rows = self._load_rows("Runs")
        return [
            row for row in rows if str(row.get("status", "")).strip().lower() == "run"
        ]

    def load_run(self, run_id):
        rows = self._load_rows("Runs")

        for row in rows:
            if str(row.get("run_id", "")).strip() == str(run_id).strip():
                return row

        raise ValueError(f"Run not found: {run_id}")
